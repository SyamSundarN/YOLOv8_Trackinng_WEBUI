"""Vehicle tracking pipeline. Heavy setup is deferred until first run_yolo() so Uvicorn/FastAPI can start immediately."""

_active_job_id = None
_run_yolo_inner = None


def run_yolo(video_path, job_id, jobs_dict, config_path=None):
    """Entry point used by FastAPI; builds the YOLO stack on first call."""
    global _run_yolo_inner
    if _run_yolo_inner is None:
        _run_yolo_inner = _bootstrap_pipeline()
    return _run_yolo_inner(video_path, job_id, jobs_dict, config_path)


def _bootstrap_pipeline():
    
    import hydra
    import torch
    import argparse
    import time
    from pathlib import Path
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    from PIL import Image, ImageChops
    
    
    import cv2
    import torch
    import torch.backends.cudnn as cudnn
    from numpy import random
    from ultralytics.yolo.engine.predictor import BasePredictor
    from ultralytics.yolo.utils import DEFAULT_CFG_PATH, ROOT, ops
    from ultralytics.yolo.utils.checks import check_imgsz
    from ultralytics.yolo.utils.plotting import Annotator, colors, save_one_box
    
    import cv2
    from deep_sort_pytorch.utils.parser import get_config
    from deep_sort_pytorch.deep_sort import DeepSort
    from collections import deque
    import numpy as np
    
    backend_dir = Path(__file__).resolve().parent
    deep_sort_cfg_path = backend_dir / "deep_sort_pytorch" / "configs" / "deep_sort_custom.yaml"
    deep_sort_custom_cfg_path = backend_dir / "deep_sort_pytorch" / "custom_config.yaml"
    excel_template_path = backend_dir / "csv_files" / "MultiDirectional_BinCountTemplate.xlsx"
    output_video_dir = backend_dir / "Output_Videos"
    traffic_counts_dir = backend_dir / "Traffic_Counts"
    output_video_dir.mkdir(exist_ok=True)
    traffic_counts_dir.mkdir(exist_ok=True)

    default_config_path = backend_dir / "csv_files" / "Douglas_Short.mp4.csv"

    def resolve_model_path():
        preferred = backend_dir / "20230420_best_weight.pt"
        if preferred.exists():
            return preferred
        candidates = sorted(backend_dir.glob("*.pt"))
        if candidates:
            return candidates[0]
        raise FileNotFoundError(
            f"YOLO weights not found in {backend_dir}. Add 20230420_best_weight.pt or another .pt file."
        )

    # Set in run_yolo so nested predict/draw code can update JOBS for the active FastAPI job.
    global coordinates
    coordinates = pd.read_csv(default_config_path, header=None)
    filepath = ""
    input_filename = ["video.mp4"]
    out = None

    fps = 0
    total_frames = 0
    #define multi-dimensional array for current count
    global current_count, veh_count, totalNB, totalSB, total_count
    totalNB = 0
    totalSB = 0
    total_traffic_count = np.zeros(10)
    ind_total_count = [0] * 10

    #Define Empty Multi-dimensional List for Count
    current_count = np.array([[0 for col in range(13)] for row in range(10)])
    total_count = np.zeros(13)
    
    palette = (2 ** 11 - 1, 2 ** 15 - 1, 2 ** 20 - 1)
    color_list = [(255, 0, 0),    # Blue
          (0, 0, 255),    # Red
          (0, 255, 0),    # Green
          (0, 165, 255),  # Orange
          (0, 255, 255),  # Yellow
          (147, 20, 255), # Pink
          (0, 215, 255),  # Gold
          (128, 0, 0),    # Navy Blue
          (238, 130, 238),# Violet
          (144, 238, 144) # Light Green
         ]

    #data_deque = {}
    deepsort = None
    totalcount = []
    


    def init_tracker():
        global deepsort
        cfg_deep = get_config()
        cfg_deep.merge_from_file(str(deep_sort_cfg_path))
        cfg_deep.DEEPSORT.REID_CKPT = str(
            (backend_dir / cfg_deep.DEEPSORT.REID_CKPT).resolve()
        )

        deepsort= DeepSort(cfg_deep.DEEPSORT.REID_CKPT,
                                max_dist=cfg_deep.DEEPSORT.MAX_DIST, min_confidence=cfg_deep.DEEPSORT.MIN_CONFIDENCE,
                                nms_max_overlap=cfg_deep.DEEPSORT.NMS_MAX_OVERLAP, max_iou_distance=cfg_deep.DEEPSORT.MAX_IOU_DISTANCE,
                                max_age=cfg_deep.DEEPSORT.MAX_AGE, n_init=cfg_deep.DEEPSORT.N_INIT, nn_budget=cfg_deep.DEEPSORT.NN_BUDGET,
                                use_cuda=torch.cuda.is_available())
    ##########################################################################################
    def xyxy_to_xywh(*xyxy):
        """" Calculates the relative bounding box from absolute pixel values. """
        bbox_left = min([xyxy[0].item(), xyxy[2].item()])
        bbox_top = min([xyxy[1].item(), xyxy[3].item()])
        bbox_w = abs(xyxy[0].item() - xyxy[2].item())
        bbox_h = abs(xyxy[1].item() - xyxy[3].item())
        x_c = (bbox_left + bbox_w / 2)
        y_c = (bbox_top + bbox_h / 2)
        w = bbox_w
        h = bbox_h
        return x_c, y_c, w, h

    def xyxy_to_tlwh(bbox_xyxy):
        tlwh_bboxs = []
        for i, box in enumerate(bbox_xyxy):
            x1, y1, x2, y2 = [int(i) for i in box]
            top = x1
            left = y1
            w = int(x2 - x1)
            h = int(y2 - y1)
            tlwh_obj = [top, left, w, h]
            tlwh_bboxs.append(tlwh_obj)
        return tlwh_bboxs

    def compute_color_for_labels(label):
        """
        Simple function that adds fixed color depending on the class
        """
        if label == 0: #person
            color = (85,45,255)
        elif label == 2: # Car
            color = (222,82,175)
        elif label == 3:  # Motobike
            color = (0, 204, 255)
        elif label == 5:  # Bus
            color = (0, 149, 255)
        else:
            color = [int((p * (label ** 2 - label + 1)) % 255) for p in palette]
        return tuple(color)

    def UI_box(x, img, color=None, label=None, line_thickness=None):
        # Plots one bounding box on image img
        tl = line_thickness or round(0.002 * (img.shape[0] + img.shape[1]) / 2) + 1  # line/font thickness
        color = color or [random.randint(0, 255) for _ in range(3)]
        c1, c2 = (int(x[0]), int(x[1])), (int(x[2]), int(x[3]))
        w, h = int(x[2])-int(x[0]), int(x[3])-int(x[1])
        cx, cy = int(x[0])+w//2, int(x[1])+h//2
        cv2.rectangle(img, c1, c2, color, thickness=tl, lineType=cv2.LINE_AA)
        if label:
            tf = max(tl - 1, 1)  # font thickness
            t_size = cv2.getTextSize(label, 0, fontScale=tl / 3, thickness=tf)[0]

            img = cv2.rectangle(img, (int(x[0]), int(x[1])), (int(x[2]), int(x[3])), color, 2)
        
            cv2.putText(img, label, (c1[0], c1[1] - 2), 0, tl / 3, [225, 255, 255], thickness=tf, lineType=cv2.LINE_AA)

    def draw_boxes(img, bbox, names,object_id, identities=None, offset=(0, 0)):
        print("frame_num: "+ str(frame_num))

        height, width, _ = img.shape
        # remove tracked point from buffer if object is lost

        for i, box in enumerate(bbox):
            x1, y1, x2, y2 = [int(i) for i in box]
            x1 += offset[0]
            x2 += offset[0]
            y1 += offset[1]
            y2 += offset[1]

            # code to find center of bottom edge
            center = (int((x2+x1)/ 2), int((y1+y2)/2))

            # get ID of object
            id = int(identities[i]) if identities is not None else 0

        
            color = compute_color_for_labels(object_id[i])
            obj_name = names[object_id[i]]
            label = obj_name

            center_x = int((box[0]+box[2])/ 2)
            center_y = int((box[1]+2*box[3])/3)
            cv2.circle(img,(center_x, center_y),5,(0,0,255),cv2.FILLED)

            maxX = np.zeros(10)
            maxY = np.zeros(10)
            minX = np.zeros(10)
            minY = np.zeros(10)

            #Counting boundary for first set
            for i in range(len(coordinates)):
                maxX[i] = max(coordinates[0][i], coordinates[2][i])
                minX[i] = min(coordinates[0][i], coordinates[2][i])
                maxY[i] = max(coordinates[1][i], coordinates[3][i])
                minY[i] = min(coordinates[1][i], coordinates[3][i])

                #Draw Lines:
                cv2.line(img, (coordinates[0][i], coordinates[1][i]), (coordinates[2][i], coordinates[3][i]), color_list[i], thickness=2)

                #Count and Record Count
                if center_y <= maxY[i] and center_y >= minY[i] and center_x <= maxX[i] and center_x >= minX[i]:
                    if totalcount.count(id) == 0:
                        totalcount.append(id)
                        for j in range(13):
                            class_name = "Class-"+str(j+1)
                            if obj_name == class_name:
                                current_count[i][j]+=1
                                if _active_job_id and _active_job_id in JOBS:
                                    JOBS[_active_job_id]["classCounts"] = {
                                        f"Class-{k+1}": int(total_count[k])
                                        for k in range(13)
                                    }
                        ind_total_count[i] += 1        
                #Display Total Count

                #cv2.putText(img, coordinates[4][i]+" Total Count: " + str(current_count.sum(axis=1)[i]), (20,50+30*i), 0, 0.75, color_list[i], 2)
                cv2.putText(img, coordinates[4][i]+" Total Count: " + str(ind_total_count[i]), (20,50+30*i), 0, 0.75, color_list[i], 2)

            UI_box(box, img, label=label, color=color, line_thickness=2)                       

        return img

    class DetectionPredictor(BasePredictor):
    
        global frame_num, active_row, nth_minute, current_count, total_count, time_hh, time_mm
        active_row = 11
        time_hh = 0
        time_mm = 0
        frame_num = 0
        def get_annotator(self, img):           
            return Annotator(img, line_width=self.args.line_thickness, example=str(self.model.names))

        def preprocess(self, img):
            global frame_num,active_row, nth_minute, totalNB, totalSB, time_mm, time_hh
        
        
            #Specify time interval in minutes:
            time_interval = max(int(coordinates[6][0]), 1)
            if frame_num % (fps*60*time_interval) == 1:
            
                #convert time to standard 12-hr format
                time_object = datetime.strptime(str(time_hh).zfill(2)+str(time_mm).zfill(2), "%H%M")
                formatted_time = time_object.strftime("%I:%M %p")

                output_excel_filename = str(
                    traffic_counts_dir / f"{input_filename[-1].split('.')[0]}_traffic_counts.xlsx"
                )
                if frame_num == 1:
                    output_excel_wb = load_workbook(excel_template_path)

                    for i in range(len(coordinates)):
                        if i ==0:
                            output_excel_ws = output_excel_wb["Sheet1"]
                            output_excel_ws.cell(row=6, column=1).value = time_interval
                            output_excel_ws.cell(row=4, column=2).value = coordinates[5][0]
                            output_excel_ws.cell(row=6, column=2).value = coordinates[7][0]
                            output_excel_ws.cell(row=8, column=2).value = coordinates[8][0]
                            output_excel_ws.cell(row=11, column=1).value = coordinates[9][0]                            
                            output_excel_wb['Sheet1'].title = coordinates[4][i]                            
                        else:
                            sheet_to_copy = output_excel_wb[coordinates[4][0]]
                            new_sheet = output_excel_wb.copy_worksheet(sheet_to_copy)
                            new_sheet.title = coordinates[4][i]
                    sheet_to_copy = output_excel_wb[coordinates[4][0]]
                    new_sheet_1 = output_excel_wb.copy_worksheet(sheet_to_copy)
                    new_sheet_1.title = "Total"     
                else:
                    output_excel_wb = load_workbook(output_excel_filename)

                for i in range(len(coordinates)):
                    output_excel_ws = output_excel_wb[coordinates[4][i]]
                    output_excel_ws.cell(row=4, column = 3).value = coordinates[4][i]
                    for j in range(3,16):
                        cellref = output_excel_ws.cell(row =active_row, column=j)
                        cellref.value = current_count[i][j-3]   
                        total_count[j-3]+= current_count[i][j-3] 
                    output_excel_ws.cell(row=active_row, column = 16).value = current_count.sum(axis=1)[i]
                    output_excel_ws.cell(row=active_row, column=1).value = coordinates[9][0] 
                    output_excel_ws.cell(row=active_row, column = 2).value = formatted_time
            
                output_excel_ws = output_excel_wb["Total"]
                for j in range(3,16):
                    cellref = output_excel_ws.cell(row =active_row, column=j)
                    cellref.value = total_count[j-3] 
                output_excel_ws.cell(row=active_row, column = 16).value = total_count.sum()
                output_excel_ws.cell(row=active_row, column=1).value = coordinates[9][0] 
                output_excel_ws.cell(row=active_row, column = 2).value = formatted_time
                output_excel_ws.cell(row=4, column = 3).value = "Total"

                output_excel_wb.save(output_excel_filename)

                active_row+=1
                if time_mm+time_interval >=60:
                    time_hh = int(time_hh + (time_mm+time_interval)/60)
                    time_mm = int((time_mm+time_interval)%60)
                else:
                    time_mm = int(time_mm+time_interval)

            
                #Reset Counters:
                current_count[:,:] = np.array([[0 for col in range(13)] for row in range(10)])
                total_count[:] = np.zeros(13)
                print('Count Printed!')
            
       
            img = torch.from_numpy(img).to(self.model.device)
            img = img.half() if self.model.fp16 else img.float()  # uint8 to fp16/32
            img /= 255  # 0 - 255 to 0.0 - 1.0
            frame_num = frame_num+1
            return img

        def postprocess(self, preds, img, orig_img):
        
            global frame_num
            preds = ops.non_max_suppression(preds,
                                            self.args.conf,
                                            self.args.iou,
                                            agnostic=self.args.agnostic_nms,
                                            max_det=self.args.max_det)

            for i, pred in enumerate(preds):
                shape = orig_img.shape
                #shape = orig_img[i].shape if self.webcam else orig_img.shape
                pred[:, :4] = ops.scale_boxes(img.shape[2:], pred[:, :4], shape).round()
        
            #Compare frames each 15 frames
            # assign initial image for reference
            global reference_frame
            if frame_num == 1:
                reference_frame = orig_img             
       
            if frame_num%(10) == 1:
            
                # read the image into a numpy array

                check_frame = np.array(orig_img)
                reference_frame = np.array(reference_frame)


                # convert the numpy array to a PIL Image object
                check_pil = Image.fromarray(np.uint8(check_frame))
                reference_pil = Image.fromarray(np.uint8(reference_frame))


                # finding difference
                diff = ImageChops.difference(reference_pil, check_pil)

                diff_array = np.asarray(diff)

                # define a threshold, 50 
                thresh = 50

                # threshold the image
                img_binary = cv2.threshold(diff_array, thresh, 255, cv2.THRESH_BINARY)[1]
                img_binary = cv2.cvtColor(img_binary, cv2.COLOR_BGR2GRAY)
                height, width = img_binary.shape

                #Count the non-zero pixels
                nzCount = cv2.countNonZero(img_binary)

                #calculate the percentage of non-zero pixels
                nzPercentage = nzCount*100/(height*width)

                print(nzPercentage)

                #Put a threshold of 20% for nzPercentage

                if nzPercentage >20:
                    print("Change of frame detected. Percentage of White Pixels:"+ str(nzPercentage))
                    raise Exception("Stopping processing of the current video")  # raise an exception to stop processing the current video
                else:
                    #update reference frame
                    reference_frame = orig_img
                
            return preds

        def write_results(self, idx, preds, batch):
        
        
            p, im, im0 = batch
            all_outputs = []
            log_string = ""
            if len(im.shape) == 3:
                im = im[None]  # expand for batch dim
            self.seen += 1
            im0 = im0.copy()
            frame = getattr(self.dataset, 'frame', 0)
            """ if self.webcam:  # batch_size >= 1
                log_string += f'{idx}: '
                frame = self.dataset.count
            else:
                frame = getattr(self.dataset, 'frame', 0) """

            self.data_path = p
            save_path = str(self.save_dir / p.name)  # im.jpg
            self.txt_path = str(self.save_dir / 'labels' / p.stem) + ('' if self.dataset.mode == 'image' else f'_{frame}')
            log_string += '%gx%g ' % im.shape[2:]  # print string
            self.annotator = self.get_annotator(im0)

            det = preds[idx]
            all_outputs.append(det)

            def push_live_frame():
                """Update preview/progress for every frame (not only when tracks exist)."""
                if not _active_job_id or _active_job_id not in JOBS:
                    return
                jid = _active_job_id
                JOBS[jid]["latestFrame"] = im0.copy()
                fr = int(getattr(self.dataset, "frame", 0))
                tf = JOBS[jid].get("totalFrames") or 0
                JOBS[jid]["currentFrame"] = fr
                if tf > 0:
                    JOBS[jid]["progress"] = min(99, int((fr / tf) * 100))

            if len(det) == 0:
                push_live_frame()
                return log_string
            for c in det[:, 5].unique():
                n = (det[:, 5] == c).sum()  # detections per class
                log_string += f"{n} {self.model.names[int(c)]}{'s' * (n > 1)}, "
            # write
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            xywh_bboxs = []
            confs = []
            oids = []
            outputs = []
            for *xyxy, conf, cls in reversed(det):
                x_c, y_c, bbox_w, bbox_h = xyxy_to_xywh(*xyxy)
                xywh_obj = [x_c, y_c, bbox_w, bbox_h]
                xywh_bboxs.append(xywh_obj)
                confs.append([conf.item()])
                oids.append(int(cls))
            xywhs = torch.Tensor(xywh_bboxs)
            confss = torch.Tensor(confs)
        
            outputs = deepsort.update(xywhs, confss, oids, im0)
            if len(outputs) > 0:
                bbox_xyxy = outputs[:, :4]
                identities = outputs[:, -2]
                object_id = outputs[:, -1]
                draw_boxes(im0, bbox_xyxy, self.model.names, object_id,identities)
            push_live_frame()
            return log_string
    

    @hydra.main(version_base=None, config_path=str(DEFAULT_CFG_PATH.parent), config_name=DEFAULT_CFG_PATH.name)
    def predict(cfg):
        
            init_tracker()
            cfg.cfg = str(deep_sort_custom_cfg_path)
            cfg.save_conf= True # save results with confidence scores
            cfg.iou = 0.8
            cfg.conf = 0.25
            cfg.model = str(resolve_model_path())
            #cfg.imgsz = check_imgsz(cfg.imgsz, min_dim=2)  # check image size
            cfg.source = filepath
            cfg.project = str(output_video_dir / f"Count_{input_filename[-1]}")
            cfg.mode = "predict"
            cfg.show = False
            predictor = DetectionPredictor(cfg)
            predictor()

    def run_yolo(video_path, job_id, jobs_dict, config_path=None):
        global JOBS, filepath, _active_job_id, coordinates, current_count, total_count, fps, active_row, time_hh, time_mm, frame_num
        JOBS = jobs_dict
        filepath = str(Path(video_path).resolve())
        _active_job_id = job_id
        input_name = Path(filepath).name
        input_filename[:] = [input_name]

        config_file = Path(config_path).resolve() if config_path else default_config_path
        if not config_file.exists():
            raise FileNotFoundError(f"Tracking config not found: {config_file}")

        coordinates = pd.read_csv(config_file, header=None)
        if len(coordinates) == 0:
            raise ValueError(f"Tracking config is empty: {config_file}")

        current_count = np.array([[0 for col in range(13)] for row in range(10)])
        total_count = np.zeros(13)
        totalcount.clear()
        for index in range(len(ind_total_count)):
            ind_total_count[index] = 0
        active_row = 11
        frame_num = 0

        time_hhmm = "".join(ch for ch in str(coordinates[10][0]) if ch.isdigit()).zfill(4)
        time_hh = int(time_hhmm[:2])
        time_mm = int(time_hhmm[2:])

        cap = cv2.VideoCapture(filepath)
        if cap.isOpened():
            JOBS[job_id]["totalFrames"] = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            JOBS[job_id]["fps"] = float(cap.get(cv2.CAP_PROP_FPS)) or 0.0
            fps = int(cap.get(cv2.CAP_PROP_FPS)) or 1
        else:
            cap.release()
            raise FileNotFoundError(f"Unable to open video for processing: {filepath}")
        cap.release()

        process_error = None
        try:
            predict()
        except Exception as e:
            print(e)
            process_error = e
        finally:
            if job_id in JOBS:
                if process_error is not None:
                    JOBS[job_id]["status"] = "error"
                    JOBS[job_id]["errorMessage"] = str(process_error)
                else:
                    JOBS[job_id]["status"] = "complete"
                    JOBS[job_id]["progress"] = 100
                    try:
                        results = {}
                        for i in range(len(coordinates)):
                            line_name = str(coordinates[4][i])
                            results[line_name] = {
                                f"Class-{k+1}": int(current_count[i][k])
                                for k in range(13)
                            }
                        JOBS[job_id]["results"] = results
                    except Exception as ex:
                        print(ex)
                        cc = JOBS[job_id].get("classCounts") or {}
                        JOBS[job_id]["results"] = {"Total": cc}
            _active_job_id = None


    return run_yolo
